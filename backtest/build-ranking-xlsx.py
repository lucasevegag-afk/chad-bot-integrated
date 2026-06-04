"""
Genera Excel completo con las 53 estrategias del registry,
ordenadas de mayor a menor TotalR 5y.
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

with open('backtest/strategies-export.json', 'r', encoding='utf-8') as f:
    strategies = json.load(f)

def detect_style(s):
    """Clasifica el estilo de operación según la config."""
    c = s['config']
    has_partial = bool(c.get('S1_PARTIAL_TP_MULT'))
    tp = float(c.get('S1_TP_MULT', 2))
    sl = float(c.get('S1_SL_MULT', 1))

    if has_partial:
        ptp = float(c['S1_PARTIAL_TP_MULT'])
        if tp >= 2.0:
            return 'Partial híbrido (scalp + intradía)'
        else:
            return 'Partial scalp-puro'

    if tp <= 0.3:
        return 'Sniper ultra-scalp'
    if tp <= 0.5:
        return 'Scalp rápido'
    if tp <= 1.0:
        return 'Intradía corto'
    if tp <= 2.0:
        return 'Intradía medio'
    return 'Swing intradía'

def detect_duration(s):
    c = s['config']
    has_partial = bool(c.get('S1_PARTIAL_TP_MULT'))
    tp = float(c.get('S1_TP_MULT', 2))
    if has_partial:
        if tp >= 2.0:
            return '3-15 min (partial) + 20-60 min (final)'
        else:
            return '3-10 min (partial) + 5-15 min (final)'
    if tp <= 0.3: return '3-8 min'
    if tp <= 0.5: return '5-15 min'
    if tp <= 1.0: return '10-30 min'
    if tp <= 2.0: return '20-60 min'
    return '30-90 min'

def detect_schedule(s):
    c = s['config']
    if c.get('S1_KILLZONES'):
        kz = c['S1_KILLZONES']
        if kz == '7,8,9,12,13,14':
            return 'London KZ (7-9 UTC) + NY KZ (12-14 UTC)'
        if kz == '8,13':
            return 'London 8 UTC + NY 13 UTC (centros)'
        return f'Killzones: {kz} UTC'
    parts = ['24h FX (sin killzones explícitas)']
    if c.get('S1_BAD_SESSIONS'):
        parts.append(f"sin {c['S1_BAD_SESSIONS']}")
    if c.get('S1_BAD_HOURS'):
        parts.append(f"sin {c['S1_BAD_HOURS']} UTC")
    if c.get('S1_BAD_DOWS'):
        dow_names = ['Dom','Lun','Mar','Mié','Jue','Vie','Sáb']
        days = [dow_names[int(d)] for d in c['S1_BAD_DOWS'].split(',')]
        parts.append(f"sin {','.join(days)}")
    return ' · '.join(parts)

def detect_origin(s):
    """Aclara de dónde viene la estrategia (swing/daytrading/scalping)."""
    c = s['config']
    has_partial = bool(c.get('S1_PARTIAL_TP_MULT'))
    tp = float(c.get('S1_TP_MULT', 2))
    category = s.get('category', 'main')

    if has_partial:
        if tp >= 2.0:
            return 'HÍBRIDO: partial cierra rápido (scalp) + final corre a TP swing/intradía. Mecanismo de gestión, NO es scalping puro.'
        else:
            return 'HÍBRIDO scalping: partial cierra ultra-rápido + final a TP scalp. Es scalping con gestión de riesgo.'
    if tp <= 0.3:
        return 'SCALPING puro: TP ultra-cercano. Trades muy rápidos.'
    if tp <= 0.5:
        return 'SCALPING / DAY TRADING corto: TP cercano, alta frecuencia.'
    if tp <= 1.0:
        return 'DAY TRADING intradía: TP moderado, trades medios.'
    if tp <= 2.0:
        return 'DAY TRADING / SWING corto: TP amplio.'
    return 'SWING INTRADÍA: TP amplio, deja correr el movimiento. NO es day-trading ni scalping clásico.'

def rr_effective(s):
    c = s['config']
    sl = float(c.get('S1_SL_MULT', 1))
    tp = float(c.get('S1_TP_MULT', 2))
    if c.get('S1_PARTIAL_TP_MULT'):
        ptp = float(c['S1_PARTIAL_TP_MULT'])
        pfrac = float(c.get('S1_PARTIAL_FRACTION', 0.5))
        # Weighted R/R: (partial_R × frac + final_R × (1-frac)) / SL
        eff = (ptp * pfrac + tp * (1 - pfrac)) / sl
        return f'1:{eff:.2f} (weighted partial)'
    return f'1:{(tp/sl):.2f}'

def breakeven_wr(s):
    c = s['config']
    sl = float(c.get('S1_SL_MULT', 1))
    tp = float(c.get('S1_TP_MULT', 2))
    if c.get('S1_PARTIAL_TP_MULT'):
        ptp = float(c['S1_PARTIAL_TP_MULT'])
        pfrac = float(c.get('S1_PARTIAL_FRACTION', 0.5))
        eff_R = (ptp * pfrac + tp * (1 - pfrac)) / sl
        return 1 / (1 + eff_R) * 100
    return 1 / (1 + tp/sl) * 100

# Ordenar por TotalR descendente
strategies.sort(key=lambda s: s['metrics']['totalR_5y'], reverse=True)

# ─── Crear workbook ───
wb = Workbook()
ws = wb.active
ws.title = 'Ranking Definitivo'

# Headers
headers = [
    '#',
    'ID',
    'Activo',
    'Tipo',
    'Estilo operativo',
    'Origen / Categoría',
    'Timeframe base',
    'Duración estimada',
    'Horario / Sesión',
    'Mecánica · Filtros',
    'SL ×ATR',
    'TP ×ATR',
    'Partial TP ×ATR',
    'Partial fracción',
    'BE after partial',
    'R/R efectivo',
    'Breakeven WR',
    'Trades 5y',
    'WinRate IS',
    'WinRate OS',
    'Margen WR sobre BE',
    'AvgR IS',
    'AvgR OS',
    'TotalR 5y',
    'Max DD (R)',
    'Max Streak Losses',
    'Decay OOS %',
    'Robustez',
    'Tagline',
    'Pros (top 3)',
    'Cons (top 3)',
    'Ideal para',
]

# Header row
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = Font(bold=True, color='FFFFFF', size=11, name='Arial')
    cell.fill = PatternFill('solid', start_color='1F2A44')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color='000000'),
                          right=Side(style='thin', color='000000'),
                          top=Side(style='thin', color='000000'),
                          bottom=Side(style='thin', color='000000'))

# Data rows
for rank, s in enumerate(strategies, start=1):
    c = s['config']
    m = s['metrics']
    expl = s['explanation']

    # Cálculos
    be_wr = breakeven_wr(s)
    margin = m['winRate_OS'] - be_wr

    pros = expl.get('pros', [])[:3]
    cons = expl.get('cons', [])[:3]
    pros_str = ' · '.join(pros)
    cons_str = ' · '.join(cons)

    row_idx = rank + 1
    row_data = [
        rank,                                       # #
        s['id'],                                    # ID
        s['asset'],                                 # Activo
        'scalping' if s.get('category')=='scalping' else 'main',  # Tipo
        detect_style(s),                            # Estilo
        detect_origin(s),                           # Origen aclaratorio
        'M5 base + multi-TF context (M15/H1/H4)',   # Timeframe
        detect_duration(s),                         # Duración
        detect_schedule(s),                         # Horario
        f"Sweep+Reclaim S1 · HTF bias filter · {expl.get('how', '')[:90]}",  # Mecánica
        float(c.get('S1_SL_MULT', 1)),              # SL
        float(c.get('S1_TP_MULT', 2)),              # TP
        float(c.get('S1_PARTIAL_TP_MULT', 0)) if c.get('S1_PARTIAL_TP_MULT') else None,
        float(c.get('S1_PARTIAL_FRACTION', 0)) if c.get('S1_PARTIAL_FRACTION') else None,
        'Sí' if c.get('S1_BE_AFTER_PARTIAL')=='1' else ('—' if not c.get('S1_PARTIAL_TP_MULT') else 'No'),
        rr_effective(s),
        be_wr / 100,                                # Breakeven WR como decimal
        m.get('trades_5y', 0),
        m['winRate_IS'] / 100,
        m['winRate_OS'] / 100,
        margin / 100,
        m['avgR_IS'],
        m['avgR_OS'],
        m['totalR_5y'],
        m['maxDD_R'],
        m['maxStreakLosses'],
        m['decay_pct'] / 100,
        s.get('robustness', 'alta'),
        s.get('tagline', ''),
        pros_str,
        cons_str,
        expl.get('idealFor', ''),
    ]
    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(vertical='center', wrap_text=True, horizontal='left')
        cell.border = Border(left=Side(style='thin', color='CCCCCC'),
                              right=Side(style='thin', color='CCCCCC'),
                              top=Side(style='thin', color='CCCCCC'),
                              bottom=Side(style='thin', color='CCCCCC'))

# ─── Formato numérico ───
# % columns: 17 (Breakeven), 19 (WR IS), 20 (WR OS), 21 (Margen), 27 (Decay)
for r in range(2, len(strategies) + 2):
    ws.cell(row=r, column=17).number_format = '0.0%'
    ws.cell(row=r, column=19).number_format = '0.0%'
    ws.cell(row=r, column=20).number_format = '0.0%'
    ws.cell(row=r, column=21).number_format = '0.0%;[Red]-0.0%'
    ws.cell(row=r, column=27).number_format = '0.0%;[Red]-0.0%'
    # R-multiples col 22, 23, 24, 25
    ws.cell(row=r, column=22).number_format = '+0.000;[Red]-0.000;-'
    ws.cell(row=r, column=23).number_format = '+0.000;[Red]-0.000;-'
    ws.cell(row=r, column=24).number_format = '+0.0"R";[Red]-0.0"R";-'
    ws.cell(row=r, column=25).number_format = '0.0"R"'
    # Trades, streak, SL, TP, partial cols - integer/decimal
    ws.cell(row=r, column=11).number_format = '0.00'  # SL
    ws.cell(row=r, column=12).number_format = '0.00'  # TP
    ws.cell(row=r, column=13).number_format = '0.00'  # Partial TP
    ws.cell(row=r, column=14).number_format = '0.00'  # Partial fraction
    ws.cell(row=r, column=18).number_format = '#,##0'  # Trades

# ─── Conditional formatting ───
# Color scale para TotalR (col 24)
totalR_range = f'X2:X{len(strategies)+1}'
ws.conditional_formatting.add(totalR_range,
    ColorScaleRule(start_type='min', start_color='F8696B',
                    mid_type='percentile', mid_value=50, mid_color='FFEB84',
                    end_type='max', end_color='63BE7B'))

# Color scale para WinRate OS (col 20)
wr_range = f'T2:T{len(strategies)+1}'
ws.conditional_formatting.add(wr_range,
    ColorScaleRule(start_type='min', start_color='F8696B',
                    mid_type='percentile', mid_value=50, mid_color='FFEB84',
                    end_type='max', end_color='63BE7B'))

# Color scale para Decay (col 27): verde si positivo (edge mejoró)
decay_range = f'AA2:AA{len(strategies)+1}'
ws.conditional_formatting.add(decay_range,
    ColorScaleRule(start_type='min', start_color='F8696B',
                    mid_type='num', mid_value=0, mid_color='FFEB84',
                    end_type='max', end_color='63BE7B'))

# Color scale para Margen WR sobre BE (col 21)
margin_range = f'U2:U{len(strategies)+1}'
ws.conditional_formatting.add(margin_range,
    ColorScaleRule(start_type='min', start_color='F8696B',
                    mid_type='num', mid_value=0, mid_color='FFEB84',
                    end_type='max', end_color='63BE7B'))

# Color scale para Max DD (col 25): rojo si alto, verde si bajo
dd_range = f'Y2:Y{len(strategies)+1}'
ws.conditional_formatting.add(dd_range,
    ColorScaleRule(start_type='min', start_color='63BE7B',
                    mid_type='percentile', mid_value=50, mid_color='FFEB84',
                    end_type='max', end_color='F8696B'))

# ─── Column widths ───
widths = {
    1: 5,    # #
    2: 22,   # ID
    3: 10,   # Activo
    4: 9,    # Tipo
    5: 28,   # Estilo
    6: 55,   # Origen (más ancho para el texto explicativo)
    7: 38,   # Timeframe
    8: 38,   # Duración
    9: 50,   # Horario
    10: 60,  # Mecánica
    11: 9,   # SL
    12: 9,   # TP
    13: 11,  # Partial TP
    14: 11,  # Partial fraction
    15: 11,  # BE after
    16: 22,  # R/R efectivo
    17: 12,  # Breakeven WR
    18: 10,  # Trades
    19: 11,  # WR IS
    20: 11,  # WR OS
    21: 13,  # Margen WR
    22: 11,  # AvgR IS
    23: 11,  # AvgR OS
    24: 12,  # TotalR
    25: 11,  # Max DD
    26: 11,  # Streak
    27: 11,  # Decay
    28: 10,  # Robustez
    29: 50,  # Tagline
    30: 60,  # Pros
    31: 60,  # Cons
    32: 60,  # Ideal para
}
for col_idx, w in widths.items():
    ws.column_dimensions[get_column_letter(col_idx)].width = w

# Row height
ws.row_dimensions[1].height = 38
for r in range(2, len(strategies) + 2):
    ws.row_dimensions[r].height = 80

# Freeze panes (first row + first 4 columns)
ws.freeze_panes = 'E2'

# ─── HOJA 2: Notas y leyenda ───
ws2 = wb.create_sheet('Notas y Leyenda')

notes = [
    ('GUÍA DE COLUMNAS · CHAD BOT STRATEGY REGISTRY', ''),
    ('', ''),
    ('Tipo', 'main = estrategia principal · scalping = trades 3-15 min con TPs ajustados'),
    ('Estilo operativo', 'Clasificación según TP/SL: Sniper, Scalp, Intradía, Swing, Partial'),
    ('', ''),
    ('━ ACLARACIÓN PARTIAL ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━', ''),
    ('Las estrategias PARTIAL no son scalp ni swing puros:', ''),
    ('  • Son HÍBRIDAS: el bot abre el trade con SL+TP ATR-based', ''),
    ('  • Al alcanzar TP1 (partial), cierra una porción (50/70%) → SL→BE', ''),
    ('  • El resto del trade corre al TP final (puede ser scalp o swing)', ''),
    ('  • Es una técnica de GESTIÓN DE RIESGO, no un estilo en sí', ''),
    ('  • Beneficio: combina alta probabilidad (TP1 cercano) + dejar correr ganadores', ''),
    ('', ''),
    ('PA1, PA3 (main)', 'Partial híbrido: TP1 scalp (+0.5 o +1.0 ATR) + final swing intradía (+2.5 ATR)'),
    ('C1, C2 (scalp)', 'Partial scalping puro: TP1 ultra-cerca (+0.2 o +0.3) + final scalp (+0.7 ATR)'),
    ('', ''),
    ('━ ESTILOS OPERATIVOS ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━', ''),
    ('Sniper ultra-scalp', 'TP×0.3 ATR (3-8 min). Wide stop SL×1.5-2 ATR. WinRate 85-91%. Ej: S8, B1, B2'),
    ('Scalp rápido', 'TP×0.5 ATR (5-15 min). SL×1 ATR. WinRate 68-75%. Ej: P1'),
    ('Intradía corto', 'TP×0.7-1.0 ATR (10-30 min). Trades medios'),
    ('Swing intradía', 'TP×2.5 ATR (30-90 min). SL×0.7 ATR. R/R 1:3.5. WinRate 25-27%. Ej: J3'),
    ('Partial híbrido', 'TP1 scalp + TP2 swing. Wr ~50-60%, captura quick profit + dejar correr'),
    ('', ''),
    ('━ MÉTRICAS CLAVE ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━', ''),
    ('Breakeven WR', '% mínimo de wins necesario para no perder, dado el R/R'),
    ('Margen WR sobre BE', 'WR actual - Breakeven WR. Más alto = más cushion'),
    ('Decay OOS %', 'Cambio en AvgR entre train (2022-24) y test (2025-26). 0% = edge perfecto. Negativo = edge se degrada'),
    ('Robustez', 'Verdict cualitativo basado en decay + DD + sample'),
    ('', ''),
    ('━ HORARIOS / SESIONES ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━', ''),
    ('London KZ', '07:00-09:59 UTC (apertura London + primera hora)'),
    ('NY KZ', '12:00-14:59 UTC (apertura NY + primera hora)'),
    ('Horas malas (10/15/18 UTC)', 'Horas con winrate <33% históricamente'),
    ('NY_PM', '17:00-19:59 UTC. Sesión perdedora cross-asset'),
    ('Sin Lunes', 'Lunes históricamente débil (noise de fin de semana)'),
    ('', ''),
    ('━ ATR & DURACIONES ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━', ''),
    ('ATR período', '14 velas M5 (≈ 70 minutos de volatilidad)'),
    ('XAU ATR típico', '~$2-3 USD'),
    ('EUR ATR típico', '~0.0005-0.0008 (5-8 pips)'),
    ('BTC ATR típico', '~$200-400'),
    ('SPX/NAS ATR', '~5-15 puntos'),
]

for r_idx, (col_a, col_b) in enumerate(notes, 1):
    a = ws2.cell(row=r_idx, column=1, value=col_a)
    b = ws2.cell(row=r_idx, column=2, value=col_b)
    a.font = Font(name='Arial', size=11, bold='━' in col_a or col_a.endswith(':') or col_a == 'GUÍA DE COLUMNAS · CHAD BOT STRATEGY REGISTRY')
    b.font = Font(name='Arial', size=10)
    a.alignment = Alignment(vertical='center', wrap_text=False)
    b.alignment = Alignment(vertical='center', wrap_text=True)
    if '━' in col_a:
        a.fill = PatternFill('solid', start_color='E8EEF7')
    if col_a == 'GUÍA DE COLUMNAS · CHAD BOT STRATEGY REGISTRY':
        a.font = Font(name='Arial', size=14, bold=True, color='1F2A44')
        ws2.row_dimensions[r_idx].height = 28

ws2.column_dimensions['A'].width = 38
ws2.column_dimensions['B'].width = 100

# ─── HOJA 3: Resumen por activo ───
ws3 = wb.create_sheet('Resumen por Activo')

# Agrupar
by_asset = {}
for s in strategies:
    a = s['asset']
    by_asset.setdefault(a, []).append(s)

# Headers
ws3.append(['Activo', 'Total Estrategias', 'Main', 'Scalping', 'Top 1 (más rentable)', 'Top TotalR (R)', 'Top WR OS', 'Top decay'])
for cell in ws3[1]:
    cell.font = Font(bold=True, color='FFFFFF', name='Arial')
    cell.fill = PatternFill('solid', start_color='1F2A44')
    cell.alignment = Alignment(horizontal='center')

for asset, slist in sorted(by_asset.items(), key=lambda x: -sum(s['metrics']['totalR_5y'] for s in x[1])):
    main_count = sum(1 for s in slist if s.get('category') != 'scalping')
    scalp_count = sum(1 for s in slist if s.get('category') == 'scalping')
    top = max(slist, key=lambda s: s['metrics']['totalR_5y'])
    ws3.append([
        asset, len(slist), main_count, scalp_count,
        top['id'], top['metrics']['totalR_5y'],
        top['metrics']['winRate_OS'] / 100,
        top['metrics']['decay_pct'] / 100,
    ])

for r in range(2, len(by_asset) + 2):
    ws3.cell(row=r, column=6).number_format = '0.0"R"'
    ws3.cell(row=r, column=7).number_format = '0.0%'
    ws3.cell(row=r, column=8).number_format = '+0.0%;[Red]-0.0%'

ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 18
ws3.column_dimensions['C'].width = 8
ws3.column_dimensions['D'].width = 10
ws3.column_dimensions['E'].width = 22
ws3.column_dimensions['F'].width = 16
ws3.column_dimensions['G'].width = 12
ws3.column_dimensions['H'].width = 14

# ─── HOJA 4: Patrones universales ───
ws4 = wb.create_sheet('Patrones Universales')

patterns_data = [
    ['ESTRATEGIA', 'CONFIG', 'ACTIVOS CONFIRMADOS', '#', 'WR RANGE', 'CATEGORÍA'],
    ['PA1 (main)', 'SL×0.7 TP×2.5 + partial 0.5/50% + BE',
     'XAU, EUR, BTC, NAS, USDCAD, USDJPY, GBPUSD, GBPAUD',
     8, '58.5% - 63.7%', 'Universal multi-asset'],
    ['S8 / B1 (main/scalp)', 'SL×1.5 TP×0.3 + killzones',
     'SPX, NAS, USDCAD, USDJPY, GBPUSD, GBPAUD, EURUSD (scalp)',
     7, '83.5% - 88.4%', 'Universal índices+FX'],
    ['B2 (scalp)', 'SL×2.0 TP×0.3',
     'XAU, SPX, NAS, USDCAD, USDJPY, GBPUSD, AUDUSD, GBPAUD',
     8, '87.4% - 91.2%', 'Universal scalp'],
    ['C2 (scalp/main)', 'SL×1.0 TP×0.7 + partial 0.3/70% + BE',
     'XAU, EUR, BTC, NAS, USDCAD, GBPUSD, GBPAUD',
     7, '76.1% - 81.0%', 'Universal partial-scalp'],
    ['J3 (main)', 'SL×0.7 TP×2.5',
     'XAU, BTC, USDCAD, USDJPY, GBPUSD, GBPAUD',
     6, '23.4% - 27.5%', 'Swing rangos amplios'],
    ['P1 (main)', 'SL×1.0 TP×0.5 + killzones',
     'XAU, EUR, BTC, NAS, USDCAD, USDJPY, GBPUSD, GBPAUD',
     8, '66.6% - 73.5%', 'Scalp killzones'],
    ['PA3 (main)', 'SL×0.7 TP×2.5 + partial 1.0/50% + BE',
     'XAU, BTC, GBPUSD, GBPAUD',
     4, '43.6% - 47.8%', 'Mid balance partial'],
]

for r_idx, row in enumerate(patterns_data, 1):
    for c_idx, val in enumerate(row, 1):
        cell = ws4.cell(row=r_idx, column=c_idx, value=val)
        if r_idx == 1:
            cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
            cell.fill = PatternFill('solid', start_color='1F2A44')
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        else:
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(vertical='center', wrap_text=True)

ws4.column_dimensions['A'].width = 22
ws4.column_dimensions['B'].width = 38
ws4.column_dimensions['C'].width = 65
ws4.column_dimensions['D'].width = 6
ws4.column_dimensions['E'].width = 16
ws4.column_dimensions['F'].width = 28
ws4.row_dimensions[1].height = 28
for r in range(2, 9):
    ws4.row_dimensions[r].height = 50

# Save
wb.save('backtest/Ranking-Estrategias-CHAD-BOT.xlsx')
print('✅ Excel generated: backtest/Ranking-Estrategias-CHAD-BOT.xlsx')
print(f'   {len(strategies)} strategies across {len(by_asset)} assets')
print(f'   4 sheets: Ranking, Notas, Resumen, Patrones')
